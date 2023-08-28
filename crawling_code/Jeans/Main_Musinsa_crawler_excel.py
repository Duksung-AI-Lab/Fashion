import cv2 as cv
from urllib.request import Request, urlopen
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import Jeans_Crop



def Musinsa_crawling(imgNum):
    global pageNum, imageNum, img_list
    # 접근할 페이지 번호
    pageNum = 1
    # 이미지 번호 (= 삽입 행 번호)
    imageNum = imgNum

    while pageNum <= 125:  ### 데이터를 읽어올 페이지 수 -> 필요한 만큼 지정해주세요
        print(pageNum)
        # 무신사스토어 데님팬츠 카테고리 & 신상품(재입고)순 링크
        url = "https://www.musinsa.com/category/003002?d_cat_cd=003002&brand=&rate=&page_kind=search&list_kind=small&sort=new&sub_sort=&page="
        url = url + str(pageNum) + "&display_cnt=90&sale_goods=&group_sale=&kids=N&ex_soldout=&color=&price1=&price2=&exclusive_yn=&shoeSizeOption=&tags=&campaign_id=&timesale_yn=&q=&includeKeywords=&measure="

        fp = urlopen(url)
        source = fp.read()
        fp.close()

        soup = BeautifulSoup(source, 'html.parser', from_encoding='utf-8')
        soup = soup.findAll('div', class_='list_img')

        # 이미지 경로를 받아 각 상품 정보 추출
        for i in soup:

            product_title = i.find("a")["title"]  # 각 상품의 title
            print(product_title)
            if product_title.find("크림") == -1 and product_title.find("화이트") == -1 and product_title.find(
                    "White") == -1 and product_title.find("white") == -1 and product_title.find("컬러") == -1 and product_title.find("GREEN") == -1 and product_title.find("옐로") == -1:
                product_url = i.find("a")["href"]  # 각 상품의 url
                print("제품 링크:", product_url)

                try:
                    src = Request(product_url, headers={'User-Agent': 'Mozilla/5.0'})  # HTTP 403 error 해결하기 위해 작성한 코드
                    src = urlopen(src).read()
                    imgsoup = BeautifulSoup(src, 'html.parser', from_encoding='utf-8')

                    img = imgsoup.find('div', {'class': 'product_img_basic'}).find('div', {'class': 'product-img'})
                    img = "https:" + img.find('img')['src']  # 상품별 상세 이미지
                    img_file_name = "Musinsa_Jeans_" + str(imageNum) + ".jpg"  # 이미지 파일명
                    img_file_path = "/hdd_ext/hdd2/halley0323/Jeans/"   # 이미지 저장 경로 지정
                    urllib.request.urlretrieve(img, img_file_path + img_file_name)

                    # 저장된 이미지 불러오기
                    image = cv.imread(img_file_path + img_file_name)
                    img_list.append(image)

                    price = imgsoup.find(id="goods_price").get_text(strip=True)
                    price = price.replace('원', '')
                    print("가격 : ", price)

                    write_ws.cell(imageNum, 1, product_title)   # 상품명
                    write_ws.cell(imageNum, 2, img_file_name)    # 이미지 파일명
                    write_ws.cell(imageNum, 7, price)  # 상품가격
                    write_ws.cell(imageNum, 8, product_url) # Link URL
                    load_wb.save("Jeans_DB.xlsx")
                    imageNum += 1
                except Exception as e:
                    print(str(e))
        pageNum += 1

if __name__ == "__main__":
    # DB 삽입 시작 행 번호
    num = 2     ### 원하는 값으로 지정

    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("Jeans_DB.xlsx", data_only=True)
    write_ws = load_wb.active

    # 크롤링한 이미지 리스트
    img_list = []
    # Excel -> name, filename, price, link
    Musinsa_crawling(num)
    # Excel -> fit, color, washing, damage
    Jeans_Crop.detect(num, img_list, write_ws, load_wb)
