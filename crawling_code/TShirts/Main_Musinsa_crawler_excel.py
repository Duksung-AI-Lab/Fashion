import cv2 as cv
from urllib.request import Request, urlopen
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import TShirts_Crop
import traceback

def Musinsa_crawling(pageNum):
    print("**************************column 1 2 3 4 *********************************")
    global img_list
    # 이미지 번호 (= 삽입 행 번호)
    imageNum = (pageNum-1)*90+1

    #while pageNum <= 1:  ### 데이터를 읽어올 페이지 수 -> 필요한 만큼 지정해주세요
        # 무신사스토어 긴팔 티셔츠 카테고리 & 신상품(재입고)순 링크
    url = "https://www.musinsa.com/category/001010?d_cat_cd=001010&brand=&rate=&page_kind=search&list_kind=small&sort=new&sub_sort=&page="
    url = url + str(pageNum) + "&display_cnt=90&sale_goods=&group_sale=&kids=N&ex_soldout=&color=&price1=&price2=&exclusive_yn=&shoeSizeOption=&tags=&campaign_id=&timesale_yn=&q=&includeKeywords=&measure="

    fp = urlopen(url)
    source = fp.read()
    fp.close()

    soup = BeautifulSoup(source, 'html.parser', from_encoding='utf-8')
    soup = soup.findAll('div', class_='list_img')

    # 이미지 경로를 받아 각 상품 정보 추출
    for i in soup:
        product_title = i.find("a")["title"]  # 각 상품의 title
        #print(product_title)
        try:
            product_url = i.find("a")["href"]  # 각 상품의 url
            #print("제품 링크:", product_url)

            src = Request(product_url, headers={'User-Agent': 'Mozilla/5.0'})  # HTTP 403 error 해결하기 위해 작성한 코드
            src = urlopen(src).read()
            imgsoup = BeautifulSoup(src, 'html.parser', from_encoding='utf-8')

            img = imgsoup.find('div', {'class': 'product_img_basic'}).find('div', {'class': 'product-img'})
            img = "https:" + img.find('img')['src']  # 상품별 상세 이미지
            img_file_name = "Musinsa_TShirt_long" + str(imageNum) + ".jpg"  # 이미지 파일명
            img_file_path = "TShirts/"   # 이미지 저장 경로 지정
            urllib.request.urlretrieve(img, img_file_path + img_file_name)
            # 저장된 이미지 불러오기
            image = cv.imread(img_file_path + img_file_name)
            img_list.append(image)

            price = imgsoup.find(id="goods_price").get_text(strip=True)
            price = price.replace('원', '')
            #print("가격 : ", price)

            write_ws.cell(imageNum, 1, product_title)   # 상품명
            write_ws.cell(imageNum, 2, img_file_name)    # 이미지 파일명
            write_ws.cell(imageNum, 3, price)  # 상품가격
            write_ws.cell(imageNum, 4, product_url) # Link URL
            load_wb.save("TShirts_DB.xlsx")
            imageNum += 1
        except Exception as e:
            print(str(e))
            print(traceback.format_exc())

if __name__ == "__main__":
    # DB 삽입 시작 행 번호
    global img_list
    #num = 4  ### 원하는 값으로 지정

    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("TShirts_DB.xlsx", data_only=True)
    write_ws = load_wb.active

    for page in range(1,75):
        # 크롤링한 이미지 리스트
        img_list = []
        # Excel -> name, filename, price, link
        Musinsa_crawling(page)
        # Excel -> printing, sleeve, neckline, color
        TShirts_Crop.detect(page, img_list, write_ws, load_wb)