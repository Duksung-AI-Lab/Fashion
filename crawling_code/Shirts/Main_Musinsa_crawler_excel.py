import os

import cv2 as cv
from urllib.request import Request, urlopen
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import Collar_Crop
import Shirts_Crop



def Musinsa_crawling(imgNum):
    global pageNum, imageNum, img_list
    # 접근할 페이지 번호
    pageNum = 1
    # 이미지 번호 (= 삽입 행 번호)
    imageNum = 1

    while pageNum <= 20:  ### 데이터를 읽어올 페이지 수 -> 필요한 만큼 지정해주세요
        # 무신사스토어 셔츠/블라우스 카테고리 & 신상품(재입고)순 링크
        url = "https://www.musinsa.com/categories/item/001002?d_cat_cd=001002&brand=&list_kind=small&sort=new&sub_sort=&page="
        url = url + str(pageNum) + "&display_cnt=90&group_sale=&exclusive_yn=&sale_goods=&timesale_yn=&ex_soldout=&kids=&color=&price1=&price2=&shoeSizeOption=&tags=&campaign_id=&includeKeywords=&measure="

        fp = urlopen(url)
        source = fp.read()
        fp.close()

        soup = BeautifulSoup(source, 'html.parser', from_encoding='utf-8')
        soup = soup.findAll('div', class_='list_img')

        # 이미지 경로를 받아 각 상품 정보 추출
        for i in soup:
            product_title = i.find("a")["title"]  # 각 상품의 title
            print(product_title)
            if product_title.find("블라우스") == -1 and product_title.find("BLOUSE") == -1 and product_title.find(
                    "blouse") == -1 and product_title.find("Blouse") == -1:
                product_url = i.find("a")["href"]  # 각 상품의 url
                product_url = "https:" + product_url
                print("제품 링크:", product_url)

                try:
                    src = Request(product_url, headers={'User-Agent': 'Mozilla/5.0'})  # HTTP 403 error 해결하기 위해 작성한 코드
                    src = urlopen(src).read()
                    imgsoup = BeautifulSoup(src, 'html.parser', from_encoding='utf-8')

                    img = imgsoup.find('div', {'class': 'product_img_basic'}).find('div', {'class': 'product-img'})
                    img = "https:" + img.find('img')['src']  # 상품별 상세 이미지
                    img_file_name = "Musinsa_Shirt_add_" + str(imageNum) + ".jpg"  # 이미지 파일명
                    img_file_path = "/hdd_ext/hdd2/halley0323/Shirts_2/"   # 이미지 저장 경로 지정
                    urllib.request.urlretrieve(img, img_file_path + img_file_name)
                    # 저장된 이미지 불러오기
                    image = cv.imread(img_file_path + img_file_name)
                    img_list.append(image)

                    price = imgsoup.find(id="goods_price").get_text(strip=True)
                    price = price.replace('원', '')
                    print("가격 : ", price)

                    write_ws.cell(imgNum, 1, product_title)   # 상품명
                    write_ws.cell(imgNum, 2, img_file_name)    # 이미지 파일명
                    write_ws.cell(imgNum, 8, price)  # 상품가격
                    write_ws.cell(imgNum, 9, product_url) # Link URL
                    load_wb.save("Shirts_DB.xlsx")
                    imageNum += 1
                    imgNum += 1
                except Exception as e:
                    print(str(e))
        pageNum += 1

if __name__ == "__main__":
    # DB 삽입 시작 행 번호
    num = 2    ### 원하는 값으로 지정

    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("Shirts_DB.xlsx", data_only=True)
    write_ws = load_wb.active

    # 크롤링한 이미지 리스트
    img_list = []
    # Excel -> name, filename, price, link
    Musinsa_crawling(num)

    # Excel -> collar
    Collar_Crop.detect(num, img_list, write_ws, load_wb)
    # Excel -> pattern, pocket, color
    Shirts_Crop.detect(num, img_list, write_ws, load_wb)